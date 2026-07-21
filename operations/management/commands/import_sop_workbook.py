from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from operations.models import (
    Equipment,
    InventoryItem,
    OperationalTask,
    Service,
    ServiceDetail,
    SubService,
    TaskEquipment,
    TaskInventory,
)


def rows(sheet):
    values = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(values)]
    for row in values:
        record = dict(zip(headers, row))
        if any(value not in (None, "") for value in record.values()):
            yield record


def text(value):
    return "" if value is None else str(value).strip()


def integer(value, default=0):
    if value in (None, ""):
        return default
    return int(value)


def decimal(value, default=Decimal("0")):
    if value in (None, ""):
        return default
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return default


def optional_decimal(value):
    return None if value in (None, "") else decimal(value)


def boolean(value):
    if isinstance(value, bool):
        return value
    return text(value).lower() in {"active", "true", "yes", "1", "y"}


def active(record):
    value = record.get("active", record.get("status", True))
    return boolean(value)


class Command(BaseCommand):
    help = "Import the external salon SOP workbook into operational master tables."

    def add_arguments(self, parser):
        parser.add_argument("workbook", nargs="?", default="data/Service_Operation_SOP_Upload_Workbook.xlsx")

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")
        book = load_workbook(path, read_only=True, data_only=True)
        required = {
            "Service Master", "Service Detail", "Sub-Service Master", "Task Master",
            "Task Inventory", "Task Equipment", "Inventory Master", "Equipment Master",
        }
        missing = required.difference(book.sheetnames)
        if missing:
            raise CommandError(f"Missing sheets: {', '.join(sorted(missing))}")

        counts = {}
        for row in rows(book["Service Master"]):
            Service.objects.update_or_create(
                code=text(row["service_code"]),
                defaults={
                    "name": text(row["service_name"]),
                    "category": text(row["service_category"]),
                    "customer_type": text(row["customer_type"]),
                    "service_type": text(row["service_type"]),
                    "standard_duration_minutes": integer(row["standard_duration_minutes"], 30),
                    "current_version": text(row["current_version"]),
                    "notes": text(row["notes"]),
                    "active": active(row),
                },
            )
        counts["services"] = Service.objects.count()

        for row in rows(book["Sub-Service Master"]):
            SubService.objects.update_or_create(
                code=text(row["sub_service_code"]),
                defaults={
                    "name": text(row["sub_service_name"]),
                    "category": text(row["category"]),
                    "standard_output_quantity": decimal(row["standard_output_quantity"], Decimal("1")),
                    "output_unit": text(row["output_unit"]),
                    "procedure_notes": text(row["procedure_notes"]),
                    "active": active(row),
                },
            )
        counts["sub-services"] = SubService.objects.count()

        for row in rows(book["Inventory Master"]):
            InventoryItem.objects.update_or_create(
                code=text(row["inventory_code"]),
                defaults={
                    "name": text(row["inventory_name"]),
                    "category": text(row["inventory_category"]),
                    "source_method": text(row["source_method"]),
                    "base_unit": text(row["base_unit"]),
                    "purchase_pack_quantity": optional_decimal(row["purchase_pack_quantity"]),
                    "purchase_pack_unit": text(row["purchase_pack_unit"]),
                    "resource_type": text(row["resource_type"]),
                    "notes": text(row["notes"]),
                    "active": active(row),
                },
            )
        counts["inventory items"] = InventoryItem.objects.count()

        for row in rows(book["Equipment Master"]):
            Equipment.objects.update_or_create(
                code=text(row["equipment_code"]),
                defaults={
                    "name": text(row["equipment_name"]),
                    "category": text(row["equipment_category"]),
                    "utility_type": text(row["utility_type"]),
                    "branch_code": text(row["branch_code"]),
                    "rated_power_kw": optional_decimal(row["rated_power_kw"]),
                    "water_litres_per_minute": optional_decimal(row["water_litres_per_minute"]),
                    "notes": text(row["notes"]),
                    "active": active(row),
                },
            )
        counts["equipment items"] = Equipment.objects.count()

        for row in rows(book["Task Master"]):
            OperationalTask.objects.update_or_create(
                external_id=text(row["task_id"]),
                defaults={
                    "sub_service": SubService.objects.get(code=text(row["sub_service_code"])),
                    "sequence": integer(row["task_sequence"], 10),
                    "activity_code": text(row["activity_code"]),
                    "name": text(row["task_name"]),
                    "role_code": text(row["role_code"]),
                    "employee_role": text(row["employee_role"]),
                    "active_labour_minutes": integer(row["active_labour_minutes"]),
                    "passive_time_minutes": integer(row["passive_time_minutes"]),
                    "required_skill": text(row["required_skill"]),
                    "allowed_staff_type": text(row["allowed_staff_type"]),
                    "default_customer_type": text(row["default_customer_type"]),
                    "is_skippable": boolean(row["is_skippable"]),
                    "notes": text(row["notes"]),
                    "active": active(row),
                },
            )
        counts["tasks"] = OperationalTask.objects.count()

        for row in rows(book["Service Detail"]):
            ServiceDetail.objects.update_or_create(
                external_id=text(row["service_detail_id"]),
                defaults={
                    "service": Service.objects.get(code=text(row["service_code"])),
                    "sub_service": SubService.objects.get(code=text(row["sub_service_code"])),
                    "sequence": integer(row["sequence"], 10),
                    "required_quantity": decimal(row["required_quantity"], Decimal("1")),
                    "mandatory": boolean(row["mandatory"]),
                    "mapping_status": text(row["mapping_status"]),
                    "notes": text(row["notes"]),
                    "active": active(row),
                },
            )
        counts["service mappings"] = ServiceDetail.objects.count()

        for row in rows(book["Task Inventory"]):
            TaskInventory.objects.update_or_create(
                external_id=text(row["task_inventory_id"]),
                defaults={
                    "task": OperationalTask.objects.get(external_id=text(row["task_id"])),
                    "service": Service.objects.get(code=text(row["service_code"])),
                    "inventory": InventoryItem.objects.get(code=text(row["inventory_code"])),
                    "standard_quantity": decimal(row["standard_quantity"]),
                    "unit": text(row["unit"]),
                    "waste_percent": decimal(row["waste_percent"]),
                    "effective_quantity": decimal(row["effective_quantity"]),
                    "mapping_status": text(row["mapping_status"]),
                    "notes": text(row["notes"] or row["mapping_note"]),
                    "active": active(row),
                },
            )
        counts["inventory mappings"] = TaskInventory.objects.count()

        for row in rows(book["Task Equipment"]):
            TaskEquipment.objects.update_or_create(
                external_id=text(row["task_equipment_id"]),
                defaults={
                    "task": OperationalTask.objects.get(external_id=text(row["task_id"])),
                    "equipment": Equipment.objects.get(code=text(row["equipment_code"])),
                    "quantity_required": decimal(row["quantity_required"], Decimal("1")),
                    "equipment_usage_minutes": integer(row["equipment_usage_minutes"]),
                    "utility_type": text(row["utility_type"]),
                    "utility_minutes": integer(row["utility_minutes"]),
                    "notes": text(row["notes"]),
                    "active": active(row),
                },
            )
        counts["equipment mappings"] = TaskEquipment.objects.count()

        summary = ", ".join(f"{name}: {count}" for name, count in counts.items())
        self.stdout.write(self.style.SUCCESS(f"Workbook imported successfully — {summary}"))
